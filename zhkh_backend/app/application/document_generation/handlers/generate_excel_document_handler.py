from datetime import datetime
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from app.application.common.interfaces.ceph import ICeph
from app.application.common.interfaces.request import IRequestHandler
from app.application.document_generation.commands.generate_excel_document_command import (
    GenerateExcelDocumentCommand,
)
from app.application.house.schemas.house_fields_schema import HouseFieldSchema
from app.config import settings
from app.domain.common.interfaces.repositories.house_repository import IHouseRepository
from app.infrastructure.containers.utils import Provide
from app.infrastructure.mediator.pipline_context import PipelineContext


class GenerateExcelDocumentHandler(IRequestHandler[GenerateExcelDocumentCommand, None]):
    def __init__(
        self,
        ceph: ICeph = Provide[ICeph],
        house_repository: IHouseRepository = Provide[IHouseRepository],
    ):
        self._ceph = ceph
        self._house_repository = house_repository

    async def handle(
        self, command: GenerateExcelDocumentCommand, context: PipelineContext
    ) -> str:
        excel_stream = await self._render_excel(command.fields)

        # Имя файла
        filename = f"houses_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        key = f"reports/{filename}"
        await self._ceph.put_object(
            bucket=settings.S3.bucket, key=key, data=excel_stream
        )

        return await self._ceph.generate_presigned_url(
            bucket=settings.S3.bucket, key=key
        )

    async def _render_excel(self, fields: List[HouseFieldSchema]) -> BytesIO:
        wb = Workbook()
        ws = wb.active

        # Добавляем заголовки из словаря, если нет, то используем имя поля
        ws.append([field.description for field in fields])

        # Центрируем заголовки
        for col_idx in range(1, len(fields) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Установка ширины столбцов
        for i, field in enumerate(fields, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = len(field.description) * 1.5

        async for chunk in self._house_repository.get_all_houses_by_chunk():
            for house in chunk:
                row = [getattr(house, f.field, "") for f in fields]
                ws.append(row)

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream
